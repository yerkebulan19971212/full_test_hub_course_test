from django import forms
from django.contrib import admin, messages
from django.shortcuts import redirect, render
from django.urls import path
from openpyxl import load_workbook

from src.common.models import School
from .models import Role, User, TokenVersion, TokenHistory, BalanceHistory

admin.site.register([
    Role, TokenVersion, TokenHistory
])


class BalanceExcelUploadForm(forms.Form):
    excel_file = forms.FileField(
        label='Excel file',
        help_text='Required columns: Аккаунт, Сумма'
    )


@admin.register(BalanceHistory)
class BalanceHistoryAdmin(admin.ModelAdmin):
    list_display = (
        'balance',
        'student',
        'id',
        'created',
    )
    search_fields = ['student__email', 'student__phone', 'student__username']
    change_list_template = 'admin/accounts/balancehistory/change_list.html'

    def get_urls(self):
        custom_urls = [
            path(
                'add-balance-excel/',
                self.admin_site.admin_view(self.add_balance_excel_view),
                name='accounts_balancehistory_add_balance_excel',
            ),
        ]
        return custom_urls + super().get_urls()

    def add_balance_excel_view(self, request):
        if request.method == 'POST':
            form = BalanceExcelUploadForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = request.FILES['excel_file']
                try:
                    wb = load_workbook(excel_file, read_only=True, data_only=True)
                    ws = wb.active

                    # Read header row
                    headers = []
                    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
                        val = str(cell.value).strip().lower() if cell.value else ''
                        headers.append(val)

                    # Map column names
                    COLUMN_MAP = {
                        'аккаунт': 'account',
                        'сумма': 'amount',
                    }

                    col_index = {}
                    for idx, header in enumerate(headers):
                        for ru_name, field_name in COLUMN_MAP.items():
                            if ru_name in header:
                                col_index[field_name] = idx
                                break

                    required = ['account', 'amount']
                    missing = [f for f in required if f not in col_index]
                    if missing:
                        messages.error(
                            request,
                            f'Missing columns: {", ".join(missing)}. '
                            f'Found headers: {", ".join(headers)}'
                        )
                        return redirect('..')

                    created = 0
                    skipped = 0
                    errors = []

                    for row_idx, row in enumerate(
                        ws.iter_rows(min_row=2, values_only=True), start=2
                    ):
                        try:
                            account = str(row[col_index['account']] or '').strip()
                            amount_val = row[col_index['amount']]

                            if not account:
                                continue

                            # Parse amount
                            if amount_val is None or (
                                isinstance(amount_val, str) and not amount_val.strip()
                            ):
                                errors.append(f'Row {row_idx}: Amount is empty')
                                continue
                            try:
                                amount = int(float(amount_val))
                            except (TypeError, ValueError):
                                errors.append(
                                    f'Row {row_idx}: Invalid amount "{amount_val}"'
                                )
                                continue
                            if amount <= 0:
                                errors.append(
                                    f'Row {row_idx}: Amount must be positive'
                                )
                                continue

                            # Search user by username
                            student = User.objects.filter(
                                username=account.lower()
                            ).first()
                            if not student:
                                errors.append(
                                    f'Row {row_idx}: User "{account}" not found'
                                )
                                continue

                            # Create BalanceHistory
                            BalanceHistory.objects.create(
                                student=student,
                                user=request.user,
                                balance=amount,
                                data=f'Excel upload row {row_idx}',
                            )
                            created += 1

                        except Exception as e:
                            errors.append(f'Row {row_idx}: {str(e)}')

                    wb.close()

                    if created:
                        messages.success(
                            request,
                            f'Successfully added balance for {created} users.'
                        )
                    if skipped:
                        messages.warning(request, f'Skipped {skipped} rows.')
                    if errors:
                        messages.error(
                            request,
                            f'Errors ({len(errors)}): ' + '; '.join(errors[:10])
                        )

                    return redirect('..')

                except Exception as e:
                    messages.error(request, f'Error processing file: {str(e)}')
                    return redirect('..')
        else:
            form = BalanceExcelUploadForm()

        context = {
            **self.admin_site.each_context(request),
            'form': form,
            'title': 'Add Balance from Excel',
            'opts': self.model._meta,
        }
        return render(
            request, 'admin/accounts/balancehistory/upload_balance.html', context
        )


class UserExcelUploadForm(forms.Form):
    excel_file = forms.FileField(
        label='Excel file',
        help_text='Required columns: Имя, Фамилия, Почта, Пароль, школа'
    )


@admin.register(User)
class UserAdmin(admin.ModelAdmin):
    list_display = (
        'username',
        'user_id',
        'id',
        'email',
        'phone',
        'login_count'
    )
    fields = [
        'username',
        'email',
        'phone',
        'role',
        'language',
        'city',
        'school',
        'balance',
        'is_google',
        'user_id',
        'login_count'
    ]
    search_fields = ['email', 'phone', 'id', 'username', 'user_id']
    change_list_template = 'admin/accounts/user/change_list.html'

    def get_urls(self):
        custom_urls = [
            path(
                'upload-excel/',
                self.admin_site.admin_view(self.upload_excel_view),
                name='accounts_user_upload_excel',
            ),
        ]
        return custom_urls + super().get_urls()

    def upload_excel_view(self, request):
        if request.method == 'POST':
            form = UserExcelUploadForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = request.FILES['excel_file']
                try:
                    wb = load_workbook(excel_file, read_only=True)
                    ws = wb.active

                    # Read header row
                    headers = []
                    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
                        val = str(cell.value).strip().lower() if cell.value else ''
                        headers.append(val)

                    # Map column names
                    COLUMN_MAP = {
                        'имя': 'first_name',
                        'фамилия': 'last_name',
                        'почта': 'email',
                        'пароль': 'password',
                        'школа': 'school',
                    }

                    col_index = {}
                    for idx, header in enumerate(headers):
                        for ru_name, field_name in COLUMN_MAP.items():
                            if ru_name in header:
                                col_index[field_name] = idx
                                break

                    required = ['first_name', 'last_name', 'email', 'password', 'school']
                    missing = [f for f in required if f not in col_index]
                    if missing:
                        messages.error(
                            request,
                            f'Missing columns: {", ".join(missing)}. '
                            f'Found headers: {", ".join(headers)}'
                        )
                        return redirect('..')

                    created = 0
                    skipped = 0
                    errors = []

                    for row_idx, row in enumerate(
                        ws.iter_rows(min_row=2, values_only=True), start=2
                    ):
                        try:
                            first_name = str(row[col_index['first_name']] or '').strip()
                            last_name = str(row[col_index['last_name']] or '').strip()
                            email = str(row[col_index['email']] or '').strip()
                            password = str(row[col_index['password']] or '').strip()
                            school_code = str(row[col_index['school']] or '').strip()

                            if not email:
                                continue

                            # Check if user already exists
                            if User.objects.filter(email=email.lower()).exists():
                                skipped += 1
                                continue

                            # Find school by name_code
                            school = None
                            if school_code:
                                school = School.objects.filter(
                                    name_code=school_code
                                ).first()
                                if not school:
                                    errors.append(
                                        f'Row {row_idx}: School "{school_code}" not found'
                                    )
                                    continue

                            user = User(
                                email=email.lower(),
                                username=email.lower(),
                                first_name=first_name,
                                last_name=last_name,
                                school=school,
                            )
                            user.set_password(password)
                            user.save()
                            created += 1

                        except Exception as e:
                            errors.append(f'Row {row_idx}: {str(e)}')

                    wb.close()

                    if created:
                        messages.success(request, f'Successfully created {created} users.')
                    if skipped:
                        messages.warning(request, f'Skipped {skipped} users (already exist).')
                    if errors:
                        messages.error(
                            request,
                            f'Errors ({len(errors)}): ' + '; '.join(errors[:10])
                        )

                    return redirect('..')

                except Exception as e:
                    messages.error(request, f'Error processing file: {str(e)}')
                    return redirect('..')
        else:
            form = UserExcelUploadForm()

        context = {
            **self.admin_site.each_context(request),
            'form': form,
            'title': 'Upload Users from Excel',
            'opts': self.model._meta,
        }
        return render(request, 'admin/accounts/user/upload_users.html', context)
